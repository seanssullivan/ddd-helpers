# -*- coding: utf-8 -*-
"""Xlsx Repository."""

# Standard Library Imports
import abc
import collections
import logging
import pathlib
from typing import Any
from typing import List
from typing import Optional
from typing import Union

# Third-Party Imports
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Local Imports
from .file_repository import AbstractFileRepository


# Initiate logger.
log = logging.getLogger(__name__)


# Constants
DEFAULT_INDEX = "id"
XLSX_EXTENSION = ".xlsx"


class AbstractXlsxRepository(AbstractFileRepository):
    """Represents an abstract Xlsx repository."""

    @property
    @abc.abstractmethod
    def columns(self) -> List[str]:
        """Column names."""
        raise NotImplementedError

    @property
    @abc.abstractmethod
    def workbook(self) -> Workbook:
        """Workbook."""
        raise NotImplementedError


class XlsxRepository(AbstractXlsxRepository):
    """Implements an Xlsx repository.

    This repository reads data from an Excel file.

    Args:
        __filepath: Path to xlsx file.
        index: Name of column to use as index.
        sheet_name (optional): Name of sheet in workbook. Default ``None``.

    Attributes:
        filepath: Path to xlsx file.
        objects: Objects in repository.
        columns: Columns names.
        workbook: Workbook.

    """

    def __init__(
        self,
        __filepath: Union[pathlib.Path, str],
        /,
        index: str = DEFAULT_INDEX,
        sheet_name: Optional[str] = None,
    ) -> None:
        if __filepath.suffix.lower() != XLSX_EXTENSION:
            message = f"{__filepath!s} is not an xlsx file"
            raise ValueError(message)

        super().__init__(__filepath)
        self._index = index
        self._objects = collections.ChainMap()

        if self._filepath.exists():
            self._workbook = self._load_workbook()
            self._worksheet = self._get_worksheet(sheet_name)
            self._load()
        else:
            self._workbook = Workbook()
            self._worksheet = self._make_worksheet(sheet_name)

    @property
    def columns(self) -> List[str]:
        """Column names."""
        try:
            keys = next(iter(self.objects)).keys()
            results = list(keys)
        except StopIteration:
            return []
        else:
            return results

    @property
    def objects(self) -> List[dict]:
        """Objects in repository."""
        return list(self._objects.values())

    @property
    def workbook(self) -> Workbook:
        """Workbook."""
        return self._workbook

    def _load_workbook(self) -> Workbook:
        """Load workbook.

        Returns:
            Workbook.

        """
        result = load_workbook(self._filepath)
        return result

    def _get_worksheet(self, name: Optional[str] = None) -> Worksheet:
        """Get worksheet.

        Args:
            name (optional): Name of worksheet. Default ``None``.

        Returns:
            Worksheet.

        """
        result = (
            self.workbook[name] if name is not None else self.workbook.active
        )
        return result

    def _make_worksheet(self, name: Optional[str] = None) -> Worksheet:
        """Make worksheet.

        Args:
            name (optional): Name of worksheet. Default ``None``.

        Returns:
            Worksheet.

        """
        result = self.workbook.active
        if name is not None:
            result.title = name

        return result

    def _load(self) -> None:
        """Load objects from xlsx file."""
        for obj in self._read_contents():
            key = obj[self._index]
            self._objects[key] = obj

        self._objects = self._objects.new_child()

    def _read_contents(self) -> List[dict]:
        """Read contents of an xlsx file.

        Returns:
            Contents of xlsx file.

        """
        data = load_rows(self._worksheet)
        results = make_objects(data[1:], data[0])
        return results

    def add(self, obj: object) -> None:
        """Add object to repository.

        Args:
            obj: Object to add to repository.

        """
        if self.can_add(obj):
            key = obj[self._index]
            self._objects[key] = obj

    def can_add(self, obj: object, /) -> bool:
        """Check whether object can be added to repository.

        Args:
            obj: Object to add to repository.

        Returns:
            Whether object can be added.

        """
        key = obj[self._index]
        result = key not in self._objects
        return result

    def get(self, ref: Union[int, str]) -> Optional[object]:
        """Get object from repository.

        Args:
            ref: Reference to object.

        Returns:
            Object.

        """
        result = self._objects.get(ref, None)
        return result

    def list(self) -> List[dict]:
        """List objects in repository.

        Returns:
            Objects in repository.

        """
        results = list(self._objects.values())
        return results

    def remove(self, obj: object) -> None:
        """Remove object from repository.

        Args:
            obj: Object to remove from repository.

        """
        key = obj[self._index]
        del self._objects[key]

    def commit(self) -> None:
        """Commit changes to repository."""
        self._update_worksheet()
        self._save_file()

    def _update_worksheet(self) -> None:
        """Update worksheet."""
        recent_objects = self._objects.maps[0]
        for obj in list(recent_objects.values()):
            row = self._make_row(obj)
            self._worksheet.append(row)

        self._objects = self._objects.new_child()

    def _make_row(self, obj: object) -> List[Any]:
        """Make row.

        Args:
            obj: Object with which to make row.

        Returns:
            Row.

        """
        if not isinstance(obj, dict):
            message = f"expected type 'dict', got {type(obj)} instead"
            raise TypeError(message)

        result = list(obj.values())
        return result

    def _save_file(self) -> None:
        """Save objects to xlsx file."""
        self._workbook.save(self.filepath)

    def rollback(self) -> None:
        """Rollback changes to repository."""
        self._objects = self._objects.parents


# ----------------------------------------------------------------------------
# Helper Functions
# ----------------------------------------------------------------------------
def load_rows(__worksheet: Worksheet, /) -> List[dict]:
    """Load rows from workbook.

    Args:
        __worksheet: Workbook from which to load rows.

    Returns:
        Rows.

    """
    results = [r for r in __worksheet.iter_rows(values_only=True) if r]
    return results


def make_objects(__rows: List[list], /, columns: list) -> dict:
    """Make objects.

    Args:
        __rows: Rows from which to make objects.
        columns: Column names to use as keys.

    Returns:
        Objects.

    """
    results = [make_object(row, columns) for row in __rows]
    return results


def make_object(__row: list, /, columns: list) -> dict:
    """Make object.

    Args:
        __row: Row from which to make object.
        columns: Column names to use as keys.

    Returns:
        Object.

    """
    result = {columns[idx]: value for idx, value in enumerate(__row)}
    return result
