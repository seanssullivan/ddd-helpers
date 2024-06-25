# -*- coding: utf-8 -*-
"""Xlsx File Wrappers."""

# Standard Library Imports
import logging
import pathlib
from typing import Any
from typing import Dict
from typing import List
from typing import Optional
from typing import Type
from typing import TypeVar

# Third-Party Imports
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Local Imports
from .base_wrappers import BaseFileWrapper


# Initiate logger.
log = logging.getLogger("dodecahedron")

# Custom types
T = TypeVar("T")

# Constants
XLSX_EXTENSION = ".xlsx"


class XlsxFileWrapper(BaseFileWrapper):
    """Implements an Xlsx file wrapper.

    Args:
        filepath: Path to xlsx file.
        columns (optional): Columns. Default ``None``.
        sheet (optional): Name of sheet in workbook. Default ``None``.

    Attributes:
        columns: Columns.
        filepath: Path to xlsx file.

    Raises:
        ValueError: when `filepath` is not a xlsx file.

    """

    def __init__(
        self,
        filepath: pathlib.Path,
        /,
        columns: Optional[List[str]] = None,
    ) -> None:
        if filepath.suffix.lower() != XLSX_EXTENSION:
            message = f"{filepath!s} is not a xlsx file"
            raise ValueError(message)

        super().__init__(filepath)
        self._columns = columns

    @property
    def columns(self) -> Optional[List[str]]:
        """Column names."""
        return self._columns

    def read(
        self,
        sheet: Optional[str] = None,
        *,
        dtype: Type[T] = dict,
    ) -> List[T]:
        """Read data from xlsx file.

        Args:
            sheet (optional): Name of worksheet. Default ``None``.
            dtype (optional): Data type of returned data. Default ``dict``.

        Returns:
            Data.

        """
        if dtype is dict:
            return self._read_each_row_as_dict(sheet)

        if dtype is list:
            return self._read_each_row_as_list(sheet)

        raise NotImplementedError

    def _read_each_row_as_dict(
        self, sheet: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Read data from CSV file as list of dictionaries.

        Args:
            sheet (optional): Name of worksheet. Default ``None``.

        Returns:
            Data.

        """
        rows = self._read_rows(sheet)
        columns, records = rows[0], rows[1:]

        if not self._columns:
            self._columns = columns

        results = [
            {columns[idx]: value for idx, value in enumerate(record)}
            for record in records
        ]
        return results

    def _read_each_row_as_list(
        self, sheet: Optional[str] = None
    ) -> List[list]:
        """Read data from CSV file as list of lists.

        Args:
            sheet (optional): Name of worksheet. Default ``None``.

        Returns:
            Data.

        """
        results = self._read_rows(sheet)
        return results

    def _read_rows(self, sheet: Optional[str] = None) -> List[list]:
        """Read rows from workbook.

        Args:
            sheet (optional): Name of worksheet. Default ``None``.

        Returns:
            Rows.

        """
        worksheet = self._load_worksheet(name=sheet)
        results = [row for row in worksheet.iter_rows(values_only=True) if row]
        return results

    def _load_worksheet(self, name: Optional[str] = None) -> Worksheet:
        """Get worksheet.

        Args:
            name (optional): Name of worksheet. Default ``None``.

        Returns:
            Worksheet.

        """
        workbook = self._load_workbook()
        result = workbook[name] if name is not None else workbook.active
        return result

    def write(self, data: list, sheet: Optional[str] = None) -> None:
        """Write data to xlsx file.

        Args:
            data: Data to write to file.
            sheet (optional): Name of worksheet. Default ``None``.

        """
        if not isinstance(data, list):
            message = f"expected type 'list', got {type(data)} instead"
            raise TypeError(message)

        if all(isinstance(_, dict) for _ in data):
            self._write_each_row_from_dict(data, sheet)

        if all(isinstance(_, list) for _ in data):
            self._write_each_row_from_list(data, sheet)

        raise NotImplementedError

    def _write_each_row_from_dict(
        self, data: List[dict], sheet: Optional[str] = None
    ) -> None:
        """Write data to xlsx file from list of dictionaries.

        Args:
            data: Data to write to xlsx file.
            sheet (optional): Name of worksheet. Default ``None``.

        """
        columns = [key for key in data[0].keys()]
        rows = [list(item.values()) for item in data]
        self._write_rows([columns, *rows], sheet)

    def _write_each_row_from_list(
        self, data: List[list], sheet: Optional[str] = None
    ) -> None:
        """Write data to xlsx file from list of lists.

        Args:
            data: Data to write to xlsx file.
            sheet (optional): Name of worksheet. Default ``None``.

        """
        self._write_rows(data, sheet)

    def _write_rows(
        self, rows: list, sheet: Optional[str] = None
    ) -> List[list]:
        """Write rows to workbook.

        Args:
            sheet (optional): Name of worksheet. Default ``None``.

        Returns:
            Rows.

        """
        workbook = self._load_workbook()
        worksheet = make_worksheet(workbook, sheet)
        for row in rows:
            worksheet.append(row)

        workbook.save(self.filepath)

    def _load_workbook(self) -> Workbook:
        """Load workbook.

        Returns:
            Workbook.

        """
        result = load_workbook(self.filepath)
        return result


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def make_worksheet(
    workbook: Workbook, name: Optional[str] = None
) -> Worksheet:
    """Make worksheet.

    Args:
        workbook: Workbook.
        name (optional): Name of worksheet. Default ``None``.

    Returns:
        Worksheet.

    """
    if name is None:
        del workbook[workbook.active.title]
        return workbook.create_sheet()
    else:
        del workbook[name]
        return workbook.create_sheet(name)
