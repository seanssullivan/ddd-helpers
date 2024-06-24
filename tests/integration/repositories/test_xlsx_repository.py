# -*- coding: utf-8 -*-

# pylint: disable=import-error
# pylint: disable=missing-function-docstring

# Standard Library Imports
import pathlib
from typing import Callable

# Third-Party Imports
from openpyxl import load_workbook

# import pytest

# Local Imports
from dodecahedron.repositories import XlsxRepository


def test_saves_an_xlsx_file(tempdir: str) -> None:
    temppath = pathlib.Path(tempdir)
    filepath = temppath / "test.xlsx"
    repo = XlsxRepository(filepath)

    obj = {"id": "1", "value": "TEST"}
    repo.add(obj)
    repo.commit()

    expected = temppath / "test.xlsx"
    assert expected.exists()


def test_loads_an_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path]
) -> None:
    rows = [["id", "value"], ["1", "TEST"]]
    filepath = make_xlsx_file("test.xlsx", rows)

    repo = XlsxRepository(filepath)
    result = repo.objects

    expected = [{"id": "1", "value": "TEST"}]
    assert result == expected


def test_adds_row_to_xlsx_file(
    make_xlsx_file: Callable[..., pathlib.Path]
) -> None:
    rows = [["id", "value"], ["1", "TEST"]]
    filepath = make_xlsx_file("test.xlsx", rows)

    repo = XlsxRepository(filepath)
    obj = {"id": "2", "value": "SUCCESS"}
    repo.add(obj)
    repo.commit()

    workbook = load_workbook(filepath, data_only=True, read_only=True)
    worksheet = workbook.active
    result = worksheet["B3"].value
    assert result == "SUCCESS"


def test_adding_rows_is_idempotent(
    make_xlsx_file: Callable[..., pathlib.Path]
) -> None:
    rows = [["id", "value"], ["1", "TEST"]]
    filepath = make_xlsx_file("test.xlsx", rows)

    repo = XlsxRepository(filepath)
    obj = {"id": "2", "value": "SUCCESS"}
    repo.add(obj)
    repo.commit()

    repo.add(obj)
    repo.commit()

    workbook = load_workbook(filepath, data_only=True, read_only=True)
    worksheet = workbook.active
    result = worksheet["B3"].value
    assert result == "SUCCESS"
